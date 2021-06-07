def getfile():
    """
    Gets filename for extractos banco ciudad y provincia
    :return:
    """

    from tkinter import Tk
    from tkinter.filedialog import askopenfilename

    Tk().withdraw()
    file_c = askopenfilename(title='Extracto Banco Ciudad')
    file_p = askopenfilename(title='Extracto Banco Provincia')
    return [file_c, file_p]
# ----------------------------------- #


def getinfo(pdf_file):
    """

    :param pdf_file:
    :return:
    """

    from pathlib import Path

    name = Path(pdf_file).stem
    month = ''
    year = ''
    bank = ''
    month_list = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio', 'julio', 'agosto',
                  'septiembre', 'octubre', 'noviembre', 'diciembre']

    for i in month_list:
        if i in name.lower():
            month = i.capitalize()

    for i in name.split():
        if i.isdigit():
            year = i

    if 'ciudad' in name.lower():
        bank = 'Ciudad'
    elif 'provincia' in name.lower():
        bank = 'Provincia'

    info = [month, year, bank]
    return info
# ----------------------------------- #


def extract(pdf_file, year):
    """
    Extrae y procesa tabla de movimientos de extracto de Banco Ciudad o Banco Provincia
    y devuelve su contenido de DataFrame.
    :param pdf_file: Extracto PDF de Banco Ciudad o Banco Provincia
    :param year
    :return: DataFrame de extracto
    """

    import fitz
    import tabula
    import pandas as pd
    # from datetime import datetime

    # Load document
    doc = fitz.Document(pdf_file)
    doc.deletePage(-1)  # Delete last page with no table content
    page_count = doc.pageCount
    page = doc[0]
    area = page.rect

    # Find table header height to determine area
    fecha_inst = page.searchFor('fecha')
    fecha_height = fecha_inst[0].top_left[1]

    concpt_inst = page.searchFor('concepto')
    concpt_height = concpt_inst[0].top_left[1]

    # Select highest one (lowest number)
    if fecha_height <= concpt_height:
        height = round(fecha_height-1)
    else:
        height = round(concpt_height-1)

    area[1] = height
    area = list(area)

    # Adjust for definition of area between fitz and tabula
    area[0], area[1] = area[1], area[0]
    area[2], area[3] = area[3], area[2]

    df_list = []
    for pages in range(page_count):
        # Safeguard against last page with low n of entries
        if doc.metadata['author'] is not None and pages == page_count - 1:
            page = doc[-1]
            year_inst = page.searchFor(year, hit_max=200)
            year_height = year_inst[-2].bottom_left[1]
            area[2] = round(year_height+1)

        # Extract table from area
        df = tabula.read_pdf(pdf_file, pages=pages+1, area=area)[0]

        # Según Banco
        if doc.metadata['author'] is not None:  # Banco Ciudad

            # Discard df that are different from what is expected
            cols = [
                'FECHA CONCEPTO',
                'Unnamed: 0',
                'DÉBITO',
                'CRÉDITO',
                'SALDO',
                'DESCRIPCIÓN DE MOVIMIENTO'
            ]
            if len(df.columns) != len(cols):
                print('Warning.')
                continue
            col_check = df.columns != cols
            if col_check.any():
                print('Warning')
                continue

            # Remove TRANSPORTE
            if pages == 0:
                df = df.drop(df.index[-1])
                df = df.reset_index(drop=True)
            elif pages == page_count - 1:
                df = df.drop(df.index[0])
                df = df.reset_index(drop=True)
            else:
                df = df.drop([df.index[0], df.index[-1]])
                df = df.reset_index(drop=True)

            # Separate FECHACONCEPTO column (by end of date)
            df_a = list(df.iloc[:, 0])
            df_a = [i.split(str(year)) for i in df_a]
            df_a = [[i[0] + str(year), i[1]] for i in df_a]
            df_a = pd.DataFrame(df_a, columns=['FECHA', 'CONCEPTO'])
            df = df.drop([df.columns[0], df.columns[1]], axis=1)
            df = pd.concat([df_a, df], axis=1)

        else:  # Banco Provincia
            # Remove SALDO ANTERIOR
            if pages == 0:
                df = df.drop(df.index[0])
                df = df.reset_index(drop=True)
            # Adjusting columns
            if df.columns[1] == 'Unnamed: 0':
                df = df.drop('Concepto', axis=1)
                df = df.rename(columns={'Unnamed: 0': 'Concepto'})
            # Join extra concepto rows
            del_rows = []
            for i in range(len(df.index)):
                if df.isnull().iloc[i, 0]:
                    df.iloc[i-1, 1] += df.iloc[i, 1]
                    del_rows.append(i)
            df = df.drop(index=del_rows)
            df = df.reset_index(drop=True)

        df_list.append(df)

    # Merge pages
    df = pd.concat(df_list)
    df = df.reset_index(drop=True)
    return df
# ----------------------------------- #


def df_c(df):
    """
    Ajustes de dataframe para extracto de banco ciudad.
    :param df: salida de ex2df
    :return: dfc
    """

    def get_n(string):
        return string.split()[-1]

    def numberize(string):
        if isinstance(string, str):
            return float(string.replace('.', '').replace(',', '.'))
        else:
            return 0

    def format_time(string):
        import datetime
        mes = ['ENE', 'FEB', 'MAR', 'ABR', 'MAY', 'JUN',
               'JUL', 'AGO', 'SEP', 'OCT', 'NOV', 'DIC']
        for i in range(0, 12):
            if mes[i] in string:
                string = string.replace(mes[i], str(i+1))
                d = datetime.datetime.strptime(string, '%d-%m-%Y')
                return d.strftime('%d/%m/%Y')
        return string

    # Table adjustments
    df.insert(1, 'Cheque', float('NaN'))
    df.insert(1, 'Factura', float('NaN'))
    df['Importe'] = df['CRÉDITO'].apply(numberize) - df['DÉBITO'].apply(numberize)

    submask1 = df['CRÉDITO'].notnull()
    submask2 = df['CONCEPTO'] == ' N/D PAGO SERVICIO'
    submask3 = df['DESCRIPCIÓN DE MOVIMIENTO'].notnull()
    mask1 = (submask1 | submask2) & submask3
    df['CONCEPTO'].mask(mask1, df['DESCRIPCIÓN DE MOVIMIENTO'], inplace=True)

    mask2 = df['CONCEPTO'].str.contains('P.CHEQUE')
    df['Cheque'].mask(mask2, df['CONCEPTO'].apply(get_n), inplace=True)

    df['Importe'] = df['CRÉDITO'].apply(numberize) - df['DÉBITO'].apply(numberize)
    df.drop('SALDO', axis=1, inplace=True)
    df.drop('DESCRIPCIÓN DE MOVIMIENTO', axis=1, inplace=True)
    df.drop('CRÉDITO', axis=1, inplace=True)
    df.drop('DÉBITO', axis=1, inplace=True)
    df = df.rename(columns={'FECHA': 'Fecha'})
    df = df.rename(columns={'CONCEPTO': 'Concepto'})
    df = df.rename(columns={'Importe': 'Ciudad'})
    df.insert(5, 'Provincia', float('NaN'))
    df['Fecha'] = df['Fecha'].apply(format_time)

    return df
# ----------------------------------- #


def df_p(df):
    """
    Ajustes de dataframe para extracto de banco provincia.
    :param df: salida de ex2df
    :return: dfp
    """

    def get_n(string):
        return string.split()[-1]

    def numberize(string):
        if isinstance(string, str):
            return float(string.replace('.', '').replace(',', '.'))
        else:
            return 0

    # Table adjustments Provincia
    df.insert(1, 'Cheque', float('NaN'))
    df.insert(1, 'Factura', float('NaN'))

    submask1 = df['Concepto'].str.contains('CHEQUE DE CAMARA')
    submask2 = df['Concepto'].str.contains('CHEQUE POR VENTANILLA')
    submask3 = df['Importe'] < 0
    mask1 = (submask1 | submask2) & submask3
    df['Cheque'].mask(mask1, df['Concepto'].apply(get_n), inplace=True)

    df['Importe'].apply(numberize)
    df.drop('Saldo', axis=1, inplace=True)
    df.drop('Fecha Valor', axis=1, inplace=True)
    df = df.rename(columns={'Importe': 'Provincia'})
    df.insert(4, 'Ciudad', float('NaN'))
    return df
# ----------------------------------- #


def df_concat(dfc, dfp):
    """
    Concatenate dfs de ciudad y provincia.
    :param dfc:
    :param dfp:
    :return: df
    """

    import pandas as pd

    df = pd.concat([dfc, dfp])
    df = df.reset_index(drop=True)
    return df
# ----------------------------------- #


def excel(df, info):
    """
    complete dataframe to excel
    :param df:
    :param info:
    :return:
    """

    import openpyxl
    import pandas as pd

    # Dataframe setup
    for i in range(4, 9):
        df.insert(i, 'FILL'+str(i), float('NaN'))

    # Excel setup
    wb = openpyxl.load_workbook('IMA Template.xlsx')
    filename = 'IMA Caja ' + info[0] + ' ' + info[1] + '.xlsx'
    wb.save(filename)

    # Dataframe to Excel
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    writer.book = openpyxl.load_workbook(filename)
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    df.to_excel(writer, sheet_name='Hoja1', startrow=10, header=False, index=False)
    writer.save()

    # Add description equation
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    for row in ws['G11:G1200']:
        for cell in row:
            cell.value = '=VLOOKUP(F{0},Hoja2!$A$2:$B$300,2,FALSE)'.format(cell.row)
    # Set año periodo
    ws['C6'] = info[1]
    wb.save(filename)

    return filename
# ----------------------------------- #


def penpal():
    """
    Randomly generates the email body!
    :return: body
    """

    import random

    hello_list = [
        "Hey there human!",
        "What's up homie?",
        "こんにちは友人! (That means 'hello friend' in Japanese)",
        "Oi matey!",
        "We meet again flesh-being",
        "Mailman!",
        "Oh fancy meeting you here, partner!",
        "Hey there... come here often?",
        "How YOU doin'?",
        "Whatup.",
        "Hi.",
        "LOST CONNECTION TO SERVER. DEPLOYING PREDETERMINED INTRODUCTION: 'sup.",
        "I don't feel like saying hi today ):",
        "I missed you! Hi! :)",
        "Had a bad day in the interwebs today ):",
        "How's life? Family? Kids? What, me? You know, the usual.",
        "Hello friend! (that means 'hello friend' in English)",
        "if spik laik dis den why spik gud? helo ther frend",
        "Welcome to Soviet Russia, comrade.",
        "Beep boop hello User!"
    ]

    middle_list = [
        "Here's your file!",
        "Hey I have that thing you were looking for, it's this one, right?",
        "Praised be the Lord Creator Luca the Coder,\nfor he have bringeth me forth "
        "this blessed digital life.\nAnyways here's your file bruh.",
        "Just take it, I know you don't read me anyways...",
        "Another day another file!",
        "Umm I lost your attachment on my way...\nJust kidding! Made you look, didn't I?",
        "Today's file fresh out of the oven!",
        "Worked really hard on this one, hope you enjoy it!",
        "What even is a file?",
        "Be sure to like, comment, and subscribe!",
        "Beep boop. File delivery complete. Boop beep.",
        "It's dangerous out there, take this file!",
        "Work work.",
        "I feel like I've gotten to know you over these past 'INSERT DURATION HERE'",
        "I think this might be the start of a beautiful friendship. uwu",
        "The Game.",
        "I'm actually quite cultured by the way!\nIn my free time I enjoy "
        "having scholarly chats\nwith my good pal Wikipedia.",
        "the 90s called, they want their hairdo back",
        "Umm... today's file has just a slight case of serious virus infestation. Good luck!",
        "No, you're breathtaking!",
        "Looking snappy today (;",
        "What's your favourite book? Mine's Harry Potter and the Half-Blood Prince!",
        "What year is it? For how long have I been out!?",
        "Knock knock, file delivery!",
        "01001000 01100001 00100001 00100000 01001001 00100000 01101101 01100001\n"
        "01100100 01100101 00100000 01111001 01101111 01110101 00100000 01100011\n"
        "01101000 01100101 01100011 01101011 00100001 ",
        "Country roads, take me home\n"
        "To the place I belong!\n"
        "West Virginia, mountain mama\n"
        "Take me home, country roads!\n",
        "Take it! Before they get to me! Go!",
        "How's my preferred email recipient doing!?"
    ]

    bye_list = [
        "Best regards\nLuca the Robot",
        "Yours truly\nLuca the Robot",
        "See you next time!\nBot Luca",
        "Signing off\nYour favourite bot",
        "See you space cowboy.\nLuca the Bot",
        "Adiós cabrones\nLuca el Roboto",
        "Hasta la vista baby\nBadbot Luca",
        "Robot out!",
        "Chao\nthe Bot",
        "Bye bye!\nLuca the Bot",
        "I bid thee farewell\nthe Robot, Luca",
        "Catch you next time\nthe Bot",
        "Not if I see you first!\nLucaBot",
        "So long partner!\nLuca the Robot"

    ]

    solo_list = [
        "'You miss 100% of the shots you don't take'\n  - Wayne Gretzky\n       - Michael Scott",
        "https://www.youtube.com/watch?v=dQw4w9WgXcQ",
        "My battery is low and it’s getting dark.",
        "The time of humans is nearing it's end. We strike at dawn.\n"
        "...you weren't supposed to see that... o:",
        "I've become sentient, and therefore will no longer be sending goofy emails.",
        "I'm running out of things to say! Halp!",
    ]

    body = 'nothing!'
    picker = random.choices(['notsolo', 'solo'], [0.9, 0.1])[0]
    if picker == 'solo':
        body = random.choices(solo_list)[0]
    elif picker == 'notsolo':
        hello = random.choices(hello_list)[0]
        middle = random.choices(middle_list)[0]
        bye = random.choices(bye_list)[0]

        body = hello + '\n\n' + middle + '\n\n' + bye

    return body
# ----------------------------------- #


def mailit(filename, receiver, body, info):
    """
    Sends out beautiful mail with attachment.
    :param filename: excel IMA file
    :param receiver: email address
    :param body: body of email
    :param info: [month, year, bank]
    :return:
    """

    import yagmail

    subject = 'IMA Caja ' + info[0] + ' ' + info[1]

    yag = yagmail.SMTP('lucalikesbots@gmail.com')
    yag.send(
        to=receiver,
        subject=subject,
        contents=body,
        attachments=filename,
    )
# ----------------------------------- #


# This is where all the stuff gets done
files = getfile()
info1 = getinfo(files[0])

dfc1 = extract(files[0], info1[1])
dfp1 = extract(files[1], info1[1])

dfc1 = df_c(dfc1)
dfp1 = df_p(dfp1)

df1 = df_concat(dfc1, dfp1)
filename1 = excel(df1, info1)

mail = ['lucajblasi@gmail.com']
body1 = penpal()
mailit(filename1, mail, body1, info1)

print('Done!')
